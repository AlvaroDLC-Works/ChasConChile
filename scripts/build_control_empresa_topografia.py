from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


try:
    ROOT = Path(__file__).resolve().parents[1]
except NameError:
    ROOT = Path.cwd()
OUTPUT = ROOT / "outputs" / "Control_Empresa_Topografia_Chile.xlsx"
MAX_ROWS = 250

PRIMARY = "1F4E5F"
SECONDARY = "2F7D7E"
SOFT = "EAF4F4"
HEADER = "D8EDED"
WARN = "FCE8B2"
BAD = "F4CCCC"
GOOD = "D9EAD3"
WHITE = "FFFFFF"
GRID = "D9E2E2"


def money_fmt():
    return '$#,##0;[Red]($#,##0);-'


def pct_fmt():
    return '0.0%;[Red](0.0%);-'


def add_title(ws, title, subtitle=None, end_col=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=PRIMARY)
    cell.font = Font(color=WHITE, bold=True, size=15)
    cell.alignment = Alignment(horizontal="center")
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        sub = ws.cell(2, 1, subtitle)
        sub.fill = PatternFill("solid", fgColor=SOFT)
        sub.font = Font(color=PRIMARY, italic=True)
        sub.alignment = Alignment(horizontal="center")


def style_sheet(ws, headers_row=1, freeze=True):
    ws.sheet_view.showGridLines = False
    if freeze:
        ws.freeze_panes = ws.cell(headers_row + 1, 1)
    thin = Side(style="thin", color=GRID)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[headers_row]:
        if cell.value:
            cell.fill = PatternFill("solid", fgColor=SECONDARY)
            cell.font = Font(color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        max_len = 10
        first_real = next((cell for cell in col if not isinstance(cell, MergedCell)), None)
        if first_real is None:
            continue
        letter = get_column_letter(first_real.column)
        for cell in col[:40]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)) + 2, 38))
        ws.column_dimensions[letter].width = max_len


def add_table(ws, name, start_row, headers, rows):
    for idx, header in enumerate(headers, 1):
        ws.cell(start_row, idx, header)
    for r_idx, row in enumerate(rows, start_row + 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(r_idx, c_idx, value)
    end_row = max(start_row + len(rows), start_row + 1)
    ref = f"A{start_row}:{ws.cell(end_row, len(headers)).coordinate}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    return end_row


def add_validation(ws, cell_range, values):
    dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def apply_number_formats(ws, money_cols=(), pct_cols=(), date_cols=(), bool_cols=()):
    for col in money_cols:
        for row in range(2, MAX_ROWS + 1):
            ws[f"{col}{row}"].number_format = money_fmt()
    for col in pct_cols:
        for row in range(2, MAX_ROWS + 1):
            ws[f"{col}{row}"].number_format = pct_fmt()
    for col in date_cols:
        for row in range(2, MAX_ROWS + 1):
            ws[f"{col}{row}"].number_format = "yyyy-mm-dd"
    for col in bool_cols:
        add_validation(ws, f"{col}2:{col}{MAX_ROWS}", ["TRUE", "FALSE"])


def set_formulas(ws, formulas_by_col):
    for row in range(2, MAX_ROWS + 1):
        for col, formula in formulas_by_col.items():
            ws[f"{col}{row}"] = formula.format(r=row)


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # Dashboard
    dash = wb.create_sheet("Dashboard")
    add_title(dash, "Control Empresa Topografia Chile", "Dashboard ejecutivo para gestion diaria", 10)
    dash["A4"] = "Indicador"
    dash["B4"] = "Valor"
    dash["D4"] = "Periodo actual"
    dash["E4"] = "=MONTH(TODAY())"
    dash["F4"] = "=YEAR(TODAY())"
    indicators = [
        ("Inversion inicial total", "=Configuracion!K2"),
        ("Monto comprado", '=SUMIF(Presupuesto_Inicial!I:I,"Comprado",Presupuesto_Inicial!G:G)'),
        ("Monto pendiente", '=SUMIF(Presupuesto_Inicial!I:I,"<>Comprado",Presupuesto_Inicial!G:G)'),
        ("Facturacion mensual", "=SUMIFS(Ingresos!J:J,Ingresos!D:D,\">=\"&DATE($F$4,$E$4,1),Ingresos!D:D,\"<\"&EDATE(DATE($F$4,$E$4,1),1))"),
        ("Gastos mensuales", "=SUMIFS(Egresos!H:H,Egresos!B:B,\">=\"&DATE($F$4,$E$4,1),Egresos!B:B,\"<\"&EDATE(DATE($F$4,$E$4,1),1))"),
        ("Impuestos estimados", "=SUMIFS(Impuestos!I:I,Impuestos!B:B,$E$4,Impuestos!C:C,$F$4)"),
        ("Utilidad neta", "=B8-B9-B10"),
        ("Recuperacion inversion", "=IFERROR(SUM(KPI!K:K)/Configuracion!K2,0)"),
        ("Proyectos activos", '=COUNTIF(Proyectos!I:I,"En ejecucion")'),
        ("Cotizaciones pendientes", '=COUNTIFS(Cotizaciones!J:J,"Enviada")'),
    ]
    for i, row in enumerate(indicators, 5):
        dash.cell(i, 1, row[0])
        dash.cell(i, 2, row[1])
    dash["D7"] = "Resumen mensual"
    dash["D8"] = "Mes"
    dash["E8"] = "Ingresos"
    dash["F8"] = "Egresos"
    dash["G8"] = "Impuestos"
    dash["H8"] = "Saldo final"
    for idx, m in enumerate(range(1, 13), 9):
        dash.cell(idx, 4, m)
        dash.cell(idx, 5, f"=SUMIFS(Flujo_Caja!D:D,Flujo_Caja!B:B,D{idx},Flujo_Caja!C:C,$F$4)")
        dash.cell(idx, 6, f"=SUMIFS(Flujo_Caja!E:E,Flujo_Caja!B:B,D{idx},Flujo_Caja!C:C,$F$4)")
        dash.cell(idx, 7, f"=SUMIFS(Flujo_Caja!F:F,Flujo_Caja!B:B,D{idx},Flujo_Caja!C:C,$F$4)")
        dash.cell(idx, 8, f"=SUMIFS(Flujo_Caja!G:G,Flujo_Caja!B:B,D{idx},Flujo_Caja!C:C,$F$4)")

    # Configuracion
    config_headers = [
        "ID_Config", "Nombre_Empresa", "RUT_Empresa", "Giro", "Fecha_Inicio", "Moneda",
        "IVA_Porcentaje", "PPM_Porcentaje", "Meta_Facturacion_Mensual", "Meta_Recuperacion_Meses",
        "Inversion_Inicial_Total", "Responsable", "Tipo_Contribuyente"
    ]
    config_rows = [[
        "CFG-001", "Control Empresa Topografia Chile", "", "Servicios de topografia, geomatica y apoyo tecnico",
        date.today(), "CLP", 0.19, 0.01, 2500000, 12, "=SUM(Presupuesto_Inicial!G:G)", "", "SpA"
    ]]
    ws = wb.create_sheet("Configuracion")
    add_table(ws, "tblConfiguracion", 1, config_headers, config_rows)
    add_validation(ws, f"F2:F{MAX_ROWS}", ["CLP", "UF", "USD"])
    add_validation(ws, f"M2:M{MAX_ROWS}", ["Persona Natural", "EIRL", "SpA"])
    apply_number_formats(ws, money_cols=["I", "K"], pct_cols=["G", "H"], date_cols=["E"])
    style_sheet(ws)

    # Presupuesto inicial
    presupuesto_headers = [
        "ID_Item", "Categoria", "Item", "Descripcion", "Cantidad", "Valor_Unitario", "Subtotal",
        "Prioridad", "Estado_Compra", "Fecha_Estimada_Compra", "Fecha_Compra_Real", "Proveedor",
        "Documento_Respaldo", "Observaciones"
    ]
    presupuesto_data = [
        ("Topografia", "GPS RTK CHCNAV i73+ o equivalente", 1, 4500000),
        ("Topografia", "Tripode, baston, accesorios", 1, 500000),
        ("Topografia", "Capacitacion y puesta en marcha", 1, 300000),
        ("Computacion", "Notebook i7 / Ryzen 7, 32 GB RAM", 1, 1300000),
        ("Computacion", "SSD externo 2 TB", 1, 120000),
        ("Computacion", "Monitor 27 pulgadas", 1, 180000),
        ("Software", "Licencia anual Civil 3D", 1, 2500000),
        ("Software", "QGIS", 1, 0),
        ("Software", "CloudCompare", 1, 0),
        ("Software", "Google Earth Pro", 1, 0),
        ("Administracion y Comercial", "Constitucion SpA", 1, 150000),
        ("Administracion y Comercial", "Dominio web + hosting 1 año", 1, 80000),
        ("Administracion y Comercial", "Pagina web corporativa basica", 1, 150000),
        ("Administracion y Comercial", "Marketing inicial", 1, 300000),
        ("Administracion y Comercial", "Fondo contingencia / seguros", 1, 500000),
    ]
    rows = [[f"PRE-{i:03d}", cat, item, "", qty, unit, f"=E{i+1}*F{i+1}", "Alta", "Pendiente", "", "", "", "", ""] for i, (cat, item, qty, unit) in enumerate(presupuesto_data, 1)]
    ws = wb.create_sheet("Presupuesto_Inicial")
    add_table(ws, "tblPresupuestoInicial", 1, presupuesto_headers, rows)
    set_formulas(ws, {"G": "=IF(OR(E{r}=\"\",F{r}=\"\"),\"\",E{r}*F{r})"})
    add_validation(ws, f"H2:H{MAX_ROWS}", ["Alta", "Media", "Baja"])
    add_validation(ws, f"I2:I{MAX_ROWS}", ["Pendiente", "Cotizado", "Comprado", "Postergado", "Eliminado"])
    apply_number_formats(ws, money_cols=["F", "G"], date_cols=["J", "K"])
    style_sheet(ws)

    # Activos
    activos_headers = [
        "ID_Activo", "Nombre_Activo", "Categoria", "Marca", "Modelo", "Numero_Serie", "Fecha_Compra",
        "Valor_Compra", "Vida_Util_Meses", "Depreciacion_Mensual", "Estado", "Ubicacion", "Responsable",
        "Documento_Factura", "Observaciones"
    ]
    ws = wb.create_sheet("Activos")
    add_table(ws, "tblActivos", 1, activos_headers, [])
    set_formulas(ws, {"J": "=IFERROR(H{r}/I{r},\"\")"})
    add_validation(ws, f"K2:K{MAX_ROWS}", ["Operativo", "En mantencion", "Dado de baja", "Vendido", "Extraviado"])
    apply_number_formats(ws, money_cols=["H", "J"], date_cols=["G"])
    style_sheet(ws)

    # Gastos mensuales seed
    gastos_headers = [
        "ID_Gasto", "Mes", "Año", "Categoria", "Concepto", "Monto_Neto", "IVA", "Monto_Total",
        "Es_Recurrente", "Estado_Pago", "Fecha_Pago", "Documento_Respaldo", "Observaciones"
    ]
    gastos_seed = [
        ("Telefonia e Internet", 30000),
        ("Hosting y servicios web", 10000),
        ("Combustible y movilizacion", 200000),
        ("Contabilidad", 80000),
        ("Imprevistos y mantencion equipos", 100000),
    ]
    current = date.today()
    rows = [[f"GME-{i:03d}", current.month, current.year, "Operacion", concepto, monto, f"=F{i+1}*Configuracion!$G$2", f"=F{i+1}+G{i+1}", True, "Pendiente", "", "", ""] for i, (concepto, monto) in enumerate(gastos_seed, 1)]
    ws = wb.create_sheet("Gastos_Mensuales")
    add_table(ws, "tblGastosMensuales", 1, gastos_headers, rows)
    set_formulas(ws, {"G": "=IF(F{r}=\"\",\"\",F{r}*Configuracion!$G$2)", "H": "=IF(F{r}=\"\",\"\",F{r}+G{r})"})
    add_validation(ws, f"J2:J{MAX_ROWS}", ["Pendiente", "Pagado", "Vencido", "No aplica"])
    apply_number_formats(ws, money_cols=["F", "G", "H"], date_cols=["K"], bool_cols=["I"])
    style_sheet(ws)

    # Clientes
    clientes_headers = [
        "ID_Cliente", "Nombre_Cliente", "Tipo_Cliente", "RUT", "Contacto", "Telefono", "Email",
        "Comuna", "Region", "Sector", "Estado", "Fuente_Contacto", "Fecha_Primer_Contacto", "Observaciones"
    ]
    ws = wb.create_sheet("Clientes")
    add_table(ws, "tblClientes", 1, clientes_headers, [])
    add_validation(ws, f"C2:C{MAX_ROWS}", ["Particular", "Arquitecto", "Constructora", "Contratista", "Inmobiliaria", "Mineria", "Ingenieria", "Municipio", "Otro"])
    add_validation(ws, f"K2:K{MAX_ROWS}", ["Prospecto", "Contactado", "Cotizado", "Cliente Activo", "Cliente Inactivo", "Perdido"])
    apply_number_formats(ws, date_cols=["M"])
    style_sheet(ws)

    # Servicios
    servicios_headers = [
        "ID_Servicio", "Nombre_Servicio", "Categoria", "Descripcion", "Precio_Minimo", "Precio_Maximo",
        "Unidad_Cobro", "Requiere_Terreno", "Requiere_Gabinete", "Documentos_Entregables", "Observaciones"
    ]
    servicios_seed = [
        ("Levantamiento topografico urbano simple", "Topografia", 300000, 800000),
        ("Levantamiento loteo/parcela", "Topografia", 500000, 1500000),
        ("Georreferenciacion de imagenes historicas", "GIS", 200000, 800000),
        ("Conversion KML/KMZ/SHP/DXF", "GIS", 100000, 500000),
        ("Cubicacion de movimiento de tierras", "Ingenieria", 300000, 2000000),
        ("Modelos digitales de terreno MDT", "Topografia", 300000, 1500000),
        ("Apoyo Oficina Tecnica", "Oficina Tecnica", 500000, 2000000),
        ("Procesamiento GIS y cartografia", "GIS", 300000, 1500000),
    ]
    rows = [[f"SER-{i:03d}", nom, cat, "", pmin, pmax, "Proyecto", True, True, "Plano, informe, archivos digitales", ""] for i, (nom, cat, pmin, pmax) in enumerate(servicios_seed, 1)]
    ws = wb.create_sheet("Servicios")
    add_table(ws, "tblServicios", 1, servicios_headers, rows)
    add_validation(ws, f"G2:G{MAX_ROWS}", ["Proyecto", "Hora", "Dia", "m2", "ha", "km", "Informe"])
    apply_number_formats(ws, money_cols=["E", "F"], bool_cols=["H", "I"])
    style_sheet(ws)

    # Generic transactional sheets
    cot_headers = [
        "ID_Cotizacion", "Fecha", "ID_Cliente", "Nombre_Proyecto", "Servicio_Principal", "Descripcion",
        "Monto_Neto", "IVA", "Monto_Total", "Estado", "Fecha_Envio", "Fecha_Respuesta",
        "Probabilidad_Cierre", "Documento_Cotizacion", "Observaciones"
    ]
    ws = wb.create_sheet("Cotizaciones")
    add_table(ws, "tblCotizaciones", 1, cot_headers, [])
    set_formulas(ws, {"H": "=IF(G{r}=\"\",\"\",G{r}*Configuracion!$G$2)", "I": "=IF(G{r}=\"\",\"\",G{r}+H{r})"})
    add_validation(ws, f"J2:J{MAX_ROWS}", ["Borrador", "Enviada", "Aprobada", "Rechazada", "Vencida"])
    apply_number_formats(ws, money_cols=["G", "H", "I"], pct_cols=["M"], date_cols=["B", "K", "L"])
    style_sheet(ws)

    proy_headers = [
        "ID_Proyecto", "ID_Cotizacion", "ID_Cliente", "Nombre_Proyecto", "Tipo_Servicio", "Fecha_Inicio",
        "Fecha_Termino_Estimada", "Fecha_Termino_Real", "Estado", "Monto_Contrato_Neto", "IVA",
        "Monto_Contrato_Total", "Costo_Estimado", "Costo_Real", "Utilidad_Estimada", "Utilidad_Real",
        "Margen_Real_Porcentaje", "Ubicacion", "Comuna", "Region", "Responsable", "Observaciones"
    ]
    ws = wb.create_sheet("Proyectos")
    add_table(ws, "tblProyectos", 1, proy_headers, [])
    set_formulas(ws, {
        "K": "=IF(J{r}=\"\",\"\",J{r}*Configuracion!$G$2)",
        "L": "=IF(J{r}=\"\",\"\",J{r}+K{r})",
        "O": "=IF(OR(J{r}=\"\",M{r}=\"\"),\"\",J{r}-M{r})",
        "P": "=IF(OR(J{r}=\"\",N{r}=\"\"),\"\",J{r}-N{r})",
        "Q": "=IFERROR(P{r}/J{r},\"\")",
    })
    add_validation(ws, f"I2:I{MAX_ROWS}", ["No iniciado", "En ejecucion", "Pausado", "Terminado", "Facturado", "Cerrado"])
    apply_number_formats(ws, money_cols=["J", "K", "L", "M", "N", "O", "P"], pct_cols=["Q"], date_cols=["F", "G", "H"])
    style_sheet(ws)

    ingresos_headers = [
        "ID_Ingreso", "ID_Proyecto", "ID_Cliente", "Fecha_Emision", "Fecha_Pago", "Tipo_Documento",
        "Numero_Documento", "Monto_Neto", "IVA", "Monto_Total", "Retencion", "PPM", "Monto_Liquido",
        "Estado_Cobro", "Documento_Respaldo", "Observaciones"
    ]
    ws = wb.create_sheet("Ingresos")
    add_table(ws, "tblIngresos", 1, ingresos_headers, [])
    set_formulas(ws, {
        "I": "=IF(H{r}=\"\",\"\",H{r}*Configuracion!$G$2)",
        "J": "=IF(H{r}=\"\",\"\",H{r}+I{r})",
        "L": "=IF(H{r}=\"\",\"\",H{r}*Configuracion!$H$2)",
        "M": "=IF(H{r}=\"\",\"\",J{r}-K{r}-L{r})",
    })
    add_validation(ws, f"F2:F{MAX_ROWS}", ["Factura afecta", "Factura exenta", "Boleta", "Recibo", "Otro"])
    add_validation(ws, f"N2:N{MAX_ROWS}", ["Pendiente", "Pagado", "Vencido", "Parcial", "Anulado"])
    apply_number_formats(ws, money_cols=["H", "I", "J", "K", "L", "M"], date_cols=["D", "E"])
    style_sheet(ws)

    egresos_headers = [
        "ID_Egreso", "Fecha", "Categoria", "Concepto", "Proveedor", "Monto_Neto", "IVA", "Monto_Total",
        "Medio_Pago", "Estado_Pago", "Documento_Respaldo", "Es_Deducible", "Observaciones"
    ]
    ws = wb.create_sheet("Egresos")
    add_table(ws, "tblEgresos", 1, egresos_headers, [])
    set_formulas(ws, {"G": "=IF(F{r}=\"\",\"\",F{r}*Configuracion!$G$2)", "H": "=IF(F{r}=\"\",\"\",F{r}+G{r})"})
    add_validation(ws, f"J2:J{MAX_ROWS}", ["Pendiente", "Pagado", "Vencido", "No aplica"])
    apply_number_formats(ws, money_cols=["F", "G", "H"], date_cols=["B"], bool_cols=["L"])
    style_sheet(ws)

    impuestos_headers = [
        "ID_Impuesto", "Mes", "Año", "IVA_Debito", "IVA_Credito", "IVA_A_Pagar", "PPM",
        "Otros_Impuestos", "Total_Impuestos", "Estado", "Fecha_Pago", "Documento_Respaldo", "Observaciones"
    ]
    rows = [[f"IMP-{m:02d}-{current.year}", m, current.year, "", "", "", "", 0, "", "Pendiente", "", "", ""] for m in range(1, 13)]
    ws = wb.create_sheet("Impuestos")
    add_table(ws, "tblImpuestos", 1, impuestos_headers, rows)
    set_formulas(ws, {
        "D": "=SUMIFS(Ingresos!I:I,Ingresos!D:D,\">=\"&DATE(C{r},B{r},1),Ingresos!D:D,\"<\"&EDATE(DATE(C{r},B{r},1),1))",
        "E": "=SUMIFS(Egresos!G:G,Egresos!B:B,\">=\"&DATE(C{r},B{r},1),Egresos!B:B,\"<\"&EDATE(DATE(C{r},B{r},1),1),Egresos!L:L,TRUE)",
        "F": "=MAX(D{r}-E{r},0)",
        "G": "=SUMIFS(Ingresos!L:L,Ingresos!D:D,\">=\"&DATE(C{r},B{r},1),Ingresos!D:D,\"<\"&EDATE(DATE(C{r},B{r},1),1))",
        "I": "=F{r}+G{r}+H{r}",
    })
    add_validation(ws, f"J2:J{MAX_ROWS}", ["Pendiente", "Declarado", "Pagado", "Observado"])
    apply_number_formats(ws, money_cols=["D", "E", "F", "G", "H", "I"], date_cols=["K"])
    style_sheet(ws)

    flujo_headers = ["ID_Flujo", "Mes", "Año", "Saldo_Inicial", "Total_Ingresos", "Total_Egresos", "Total_Impuestos", "Saldo_Final", "Observaciones"]
    rows = [[f"FLU-{m:02d}-{current.year}", m, current.year, 0, "", "", "", "", ""] for m in range(1, 13)]
    ws = wb.create_sheet("Flujo_Caja")
    add_table(ws, "tblFlujoCaja", 1, flujo_headers, rows)
    set_formulas(ws, {
        "E": "=SUMIFS(Ingresos!J:J,Ingresos!D:D,\">=\"&DATE(C{r},B{r},1),Ingresos!D:D,\"<\"&EDATE(DATE(C{r},B{r},1),1))",
        "F": "=SUMIFS(Egresos!H:H,Egresos!B:B,\">=\"&DATE(C{r},B{r},1),Egresos!B:B,\"<\"&EDATE(DATE(C{r},B{r},1),1))+SUMIFS(Gastos_Mensuales!H:H,Gastos_Mensuales!B:B,B{r},Gastos_Mensuales!C:C,C{r})",
        "G": "=SUMIFS(Impuestos!I:I,Impuestos!B:B,B{r},Impuestos!C:C,C{r})",
        "H": "=D{r}+E{r}-F{r}-G{r}",
    })
    apply_number_formats(ws, money_cols=["D", "E", "F", "G", "H"])
    style_sheet(ws)

    roadmap_headers = [
        "ID_Tarea", "Etapa", "Tarea", "Descripcion", "Prioridad", "Responsable", "Fecha_Inicio",
        "Fecha_Limite", "Fecha_Cierre", "Estado", "Porcentaje_Avance", "Dependencia", "Documento_Asociado", "Observaciones"
    ]
    roadmap_seed = [
        ("Formalizacion", "Definir tipo de contribuyente y nombre comercial"),
        ("Compra de equipos", "Cotizar GPS RTK y accesorios"),
        ("Imagen comercial", "Crear dominio, correo y pagina web basica"),
        ("Captacion de clientes", "Armar base inicial de prospectos"),
        ("Primeros proyectos", "Levantar primeras cotizaciones pagadas"),
        ("Control financiero", "Registrar ingresos, egresos e impuestos mensuales"),
        ("Regularizacion tributaria", "Validar flujo mensual con contador"),
        ("Consolidacion urbana", "Estandarizar plantillas de entrega tecnica"),
        ("Entrada a mineria", "Diseñar modulo futuro de faenas, HSE y produccion"),
        ("Evolucion consultora", "Definir cartera de ingenieria, GIS, QA/QC y BIM"),
    ]
    rows = [[f"TAR-{i:03d}", etapa, tarea, "", "Alta" if i <= 4 else "Media", "", "", "", "", "Pendiente", 0, "", "", ""] for i, (etapa, tarea) in enumerate(roadmap_seed, 1)]
    ws = wb.create_sheet("Roadmap")
    add_table(ws, "tblRoadmap", 1, roadmap_headers, rows)
    add_validation(ws, f"B2:B{MAX_ROWS}", ["Formalizacion", "Compra de equipos", "Imagen comercial", "Captacion de clientes", "Primeros proyectos", "Control financiero", "Regularizacion tributaria", "Consolidacion urbana", "Entrada a mineria", "Evolucion consultora"])
    add_validation(ws, f"E2:E{MAX_ROWS}", ["Alta", "Media", "Baja"])
    add_validation(ws, f"J2:J{MAX_ROWS}", ["Pendiente", "En proceso", "Bloqueada", "Completada", "Cancelada"])
    apply_number_formats(ws, pct_cols=["K"], date_cols=["G", "H", "I"])
    style_sheet(ws)

    documentos_headers = [
        "ID_Documento", "Categoria", "Nombre_Documento", "Fecha", "Relacionado_A", "ID_Relacionado",
        "Archivo", "Vencimiento", "Estado", "Observaciones"
    ]
    ws = wb.create_sheet("Documentos")
    add_table(ws, "tblDocumentos", 1, documentos_headers, [])
    add_validation(ws, f"B2:B{MAX_ROWS}", ["Legal", "Tributario", "Comercial", "Tecnico", "Contrato", "Cotizacion", "Factura", "Certificado", "Seguro", "Otro"])
    add_validation(ws, f"E2:E{MAX_ROWS}", ["Cliente", "Cotizacion", "Proyecto", "Ingreso", "Egreso", "Activo", "Empresa", "Otro"])
    add_validation(ws, f"I2:I{MAX_ROWS}", ["Vigente", "Por vencer", "Vencido", "Archivado"])
    apply_number_formats(ws, date_cols=["D", "H"])
    ws.conditional_formatting.add(f"H2:H{MAX_ROWS}", FormulaRule(formula=[f'AND(H2<>"",H2<=TODAY()+30,H2>=TODAY())'], fill=PatternFill("solid", fgColor=WARN)))
    ws.conditional_formatting.add(f"H2:H{MAX_ROWS}", FormulaRule(formula=[f'AND(H2<>"",H2<TODAY())'], fill=PatternFill("solid", fgColor=BAD)))
    style_sheet(ws)

    kpi_headers = [
        "ID_KPI", "Mes", "Año", "Facturacion_Total", "Gastos_Totales", "Utilidad_Neta", "Margen_Neto",
        "Cotizaciones_Enviadas", "Cotizaciones_Aprobadas", "Tasa_Conversion", "Inversion_Recuperada",
        "Porcentaje_Recuperacion_Inversion", "Saldo_Caja", "Observaciones"
    ]
    rows = [[f"KPI-{m:02d}-{current.year}", m, current.year, "", "", "", "", "", "", "", "", "", "", ""] for m in range(1, 13)]
    ws = wb.create_sheet("KPI")
    add_table(ws, "tblKPI", 1, kpi_headers, rows)
    set_formulas(ws, {
        "D": "=SUMIFS(Ingresos!J:J,Ingresos!D:D,\">=\"&DATE(C{r},B{r},1),Ingresos!D:D,\"<\"&EDATE(DATE(C{r},B{r},1),1))",
        "E": "=SUMIFS(Egresos!H:H,Egresos!B:B,\">=\"&DATE(C{r},B{r},1),Egresos!B:B,\"<\"&EDATE(DATE(C{r},B{r},1),1))+SUMIFS(Gastos_Mensuales!H:H,Gastos_Mensuales!B:B,B{r},Gastos_Mensuales!C:C,C{r})+SUMIFS(Impuestos!I:I,Impuestos!B:B,B{r},Impuestos!C:C,C{r})",
        "F": "=D{r}-E{r}",
        "G": "=IFERROR(F{r}/D{r},0)",
        "H": "=COUNTIFS(Cotizaciones!J:J,\"Enviada\",Cotizaciones!B:B,\">=\"&DATE(C{r},B{r},1),Cotizaciones!B:B,\"<\"&EDATE(DATE(C{r},B{r},1),1))",
        "I": "=COUNTIFS(Cotizaciones!J:J,\"Aprobada\",Cotizaciones!B:B,\">=\"&DATE(C{r},B{r},1),Cotizaciones!B:B,\"<\"&EDATE(DATE(C{r},B{r},1),1))",
        "J": "=IFERROR(I{r}/H{r},0)",
        "K": "=MAX(0,SUM($F$2:F{r}))",
        "L": "=IFERROR(K{r}/Configuracion!$K$2,0)",
        "M": "=SUMIFS(Flujo_Caja!H:H,Flujo_Caja!B:B,B{r},Flujo_Caja!C:C,C{r})",
    })
    apply_number_formats(ws, money_cols=["D", "E", "F", "K", "M"], pct_cols=["G", "J", "L"])
    style_sheet(ws)

    # Helper documentation sheet
    schema = wb.create_sheet("Schema_AppSheet")
    schema_headers = ["Tabla", "Columna", "Tipo AppSheet", "Key", "Label", "Formula / App formula", "Validacion / Valores", "Notas"]
    schema_rows = [
        ("Configuracion", "ID_Config", "Text", "Y", "", 'UNIQUEID("CFG")', "", "Una fila principal de parametros."),
        ("Presupuesto_Inicial", "ID_Item", "Text", "Y", "", 'UNIQUEID("PRE")', "", "Tabla de inversion inicial."),
        ("Presupuesto_Inicial", "Subtotal", "Price", "", "", "[Cantidad] * [Valor_Unitario]", "", "App formula o spreadsheet formula."),
        ("Activos", "ID_Activo", "Text", "Y", "", 'UNIQUEID("ACT")', "", "Activos comprados."),
        ("Activos", "Depreciacion_Mensual", "Price", "", "", "[Valor_Compra] / [Vida_Util_Meses]", "", "Proteger contra division por cero."),
        ("Clientes", "ID_Cliente", "Text", "Y", "Y", 'UNIQUEID("CLI")', "", "CRM simple."),
        ("Cotizaciones", "ID_Cotizacion", "Text", "Y", "", 'UNIQUEID("COT")', "", "Kanban por estado."),
        ("Cotizaciones", "IVA", "Price", "", "", "[Monto_Neto] * ANY(Configuracion[IVA_Porcentaje])", "", "IVA Chile default 19%."),
        ("Proyectos", "ID_Proyecto", "Text", "Y", "Y", 'UNIQUEID("PRO")', "", "Se crea desde cotizacion aprobada."),
        ("Ingresos", "PPM", "Price", "", "", "[Monto_Neto] * ANY(Configuracion[PPM_Porcentaje])", "", "Editable desde Configuracion."),
        ("Documentos", "Archivo", "File", "", "", "", "", "Adjunto en Google Drive/AppSheet."),
        ("Roadmap", "Porcentaje_Avance", "Percent", "", "", "", "0 a 100%", "Checklist de consolidacion."),
    ]
    add_table(schema, "tblSchemaAppSheet", 1, schema_headers, schema_rows)
    style_sheet(schema)

    # Listas
    listas = wb.create_sheet("Listas")
    lists = {
        "Estado_Compra": ["Pendiente", "Cotizado", "Comprado", "Postergado", "Eliminado"],
        "Estado_Cotizacion": ["Borrador", "Enviada", "Aprobada", "Rechazada", "Vencida"],
        "Estado_Proyecto": ["No iniciado", "En ejecucion", "Pausado", "Terminado", "Facturado", "Cerrado"],
        "Estado_Pago_Cobro": ["Pendiente", "Pagado", "Vencido", "Parcial", "Anulado", "No aplica"],
        "Categoria_Documento": ["Legal", "Tributario", "Comercial", "Tecnico", "Contrato", "Cotizacion", "Factura", "Certificado", "Seguro", "Otro"],
        "Etapas_Roadmap": ["Formalizacion", "Compra de equipos", "Imagen comercial", "Captacion de clientes", "Primeros proyectos", "Control financiero", "Regularizacion tributaria", "Consolidacion urbana", "Entrada a mineria", "Evolucion consultora"],
    }
    for c, (title, vals) in enumerate(lists.items(), 1):
        listas.cell(1, c, title)
        for r, val in enumerate(vals, 2):
            listas.cell(r, c, val)
    style_sheet(listas)

    # Styling dashboard last
    style_sheet(dash, headers_row=4, freeze=False)
    for cell in ["B5", "B6", "B7", "B8", "B9", "B10", "B11"]:
        dash[cell].number_format = money_fmt()
    dash["B12"].number_format = pct_fmt()
    for row in range(9, 21):
        for col in range(5, 9):
            dash.cell(row, col).number_format = money_fmt()
    for row in dash["A5:B14"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=SOFT)
    for cell in dash["D7:H7"][0]:
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.font = Font(color=WHITE, bold=True)

    bar = BarChart()
    bar.title = "Ingresos, egresos e impuestos"
    data = Reference(dash, min_col=5, max_col=7, min_row=8, max_row=20)
    cats = Reference(dash, min_col=4, min_row=9, max_row=20)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.y_axis.numFmt = '$#,##0'
    dash.add_chart(bar, "J5")

    line = LineChart()
    line.title = "Saldo final"
    data = Reference(dash, min_col=8, min_row=8, max_row=20)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.y_axis.numFmt = '$#,##0'
    dash.add_chart(line, "J21")

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)

    # Verification pass: load formulas and check sheets exist, key seeds and formulas.
    check = load_workbook(OUTPUT, data_only=False)
    required = {
        "Configuracion", "Presupuesto_Inicial", "Activos", "Gastos_Mensuales", "Clientes",
        "Servicios", "Cotizaciones", "Proyectos", "Ingresos", "Egresos", "Impuestos",
        "Flujo_Caja", "Roadmap", "Documentos", "KPI", "Dashboard", "Schema_AppSheet"
    }
    missing = required - set(check.sheetnames)
    if missing:
        raise RuntimeError(f"Missing sheets: {sorted(missing)}")
    total_formula = check["Configuracion"]["K2"].value
    subtotal_formula = check["Presupuesto_Inicial"]["G2"].value
    if not str(total_formula).startswith("=") or not str(subtotal_formula).startswith("="):
        raise RuntimeError("Expected formulas were not written")
    seed_total = sum(row[3] * row[2] for row in presupuesto_data)
    if seed_total != 10580000:
        raise RuntimeError(f"Seed total mismatch: {seed_total}")
    print(f"Workbook created: {OUTPUT}")
    print("Verified sheets:", len(check.sheetnames))
    print("Initial investment seed total:", seed_total)


if __name__ == "__main__":
    build()
